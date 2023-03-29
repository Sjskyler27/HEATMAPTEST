import openpyxl
import random
import os
import threading
debug = True

# print(os.getcwd())  # prints the current working directory

class Bot:
    def __init__(self, name, ws, position=0, timetocorner=0, currentcorner=0, speedofcorner=0):
        self.name = name
        self.ws = ws
        self.position = position
        self.timetocorner = timetocorner
        self.currentcorner = currentcorner
        self.speedofcorner = speedofcorner
        self.legandLine = 1
        self.drawpile = []
        self.discardpile = [(2,15),(1,13),(1,14),(3,18),(0,10),(2,16),(2,17),(1,12),(3,19),(0,11)]
        self.card = None
        self.finish = False
        self.turnstaken = 0
        self.doublecorner = False
        self.wassaveneeded= False
        self.shuffleDiscard()



    def getMapValues(self, ws):
        self.worksheet = ws

        try:
            print("getting map values:") if debug else None
            print(f"position = {self.position}") if debug else None
            # Get the value in the specified row and column
            for i in range(2,6):      
                row_num = self.position + 2
                col_num = i
                #get the actual value!!
                value = ws.cell(row=row_num, column=col_num).value
                # if isinstance(value, str) and value.startswith('='):
                #     # If the cell contains a formula, get the computed value
                #     value = ws.cell(row=row_num, column=col_num).value
                # print(value) 
                # #value = cell.value.value

                if value == 'FINISH':
                    self.finish = True
                    print('black has finished') if debug else None
                else:
                    if(i == 2):
                        self.timetocorner = int(value)
                        print(f"ttc {value}") if debug else None
                    if(i == 3):
                        if(value):
                            self.speedofcorner = int(value)
                            print(f"speed of corner {value}") if debug else None
                    if(i == 4):
                        if(value):
                            self.legandLine = int(value)
                            print(f"Legand line {value}") if debug else None
                    if(i == 5):
                            self.doublecorner = int(value)
                            print(f"Double corner danger value: {value}\n") if debug else None
                
        except FileNotFoundError:
            print("File not found.") if debug else None
        except Exception as e:
            print(f"An error occurred: {e}") if debug else None



    def drawCard(self):
        if not self.drawpile:
            self.shuffleDiscard()

        self.card = self.drawpile.pop(0)
        self.discardpile.append(self.card)
        print(f"{self.name} has drawn {self.card}") if debug else None
        


    def shuffleDiscard(self):
        print("shufflediscard") if debug else None
        # Shuffle the discard pile
        random.shuffle(self.discardpile)
        print(f"this is new order for draw{self.discardpile} \n") if debug else None

        #Use the shuffled discard pile as the new deck
        self.drawpile = self.discardpile.copy()

        # Clear the discard pile
        self.discardpile.clear()



    def move(self):
        print(f"\n{self.name} is moving from {self.position} ttc {self.timetocorner}") if debug else None

        #Check for double move, this is the most dificult part
        if(not self.doublecorner == False):
            self.wassaveneeded = True
            
            print(f"{self.name} is in danger of a double corner. Check to see how they move") if debug else None
            #check if the card they pull is actually dangerous
            #card first value < worksheet double Corner Value then safe move value + corner speed
            if(self.card[0] < self.doublecorner):
                self.position = self.position + self.card[0] + self.speedofcorner
                print(f"by cornering {self.card[0]+ self.speedofcorner}") if debug else None
            #else card value > worksheet double Corner value not safe, move to default of the next corner 
            #but how
            #maybe position+= ttc+1 that gets it past the corner then call get map 
            # #values so you have the new ttc then do a default move. 
            else:
                print(f"by special default {self.card[0]}") if debug else None
                self.position = self.position + self.timetocorner + 1
                print('getting new ttc') if debug else None
                self.getMapValues(self.ws)
                print(f'new ttc: {self.timetocorner}') if debug else None
                self.position = self.position + self.timetocorner - self.card[0]


        else:
            if(self.legandLine != 1): #2 = false needs to be a number in the spread sheet for some reason
                #big move if behind legand line and can move full number
                if self.timetocorner - self.card[1] > 0:
                    self.position = self.position + self.card[1]
                    print(f"using big move {self.card[1]}") if debug else None
                else:
                    #move to default square if not enough spaces for big move
                    self.position = self.position + self.timetocorner - self.card[0]
                    print(f"using defaulted move of {self.card[0]}") if debug else None
            else:
                #corner if in front of legand line, corner speed plus first num
                self.position = self.position + self.card[0] + self.speedofcorner
                print(f"by cornering {self.card[0]+ self.speedofcorner}") if debug else None

        print(f"to position {self.position}") if debug else None
        self.turnstaken = self.turnstaken + 1
        print(f'turns taken: {self.turnstaken}\n') if debug else None

class Race:
    def __init__(self, track_file, sheet, times):
        self.track_file = track_file
        self.sheet = sheet
        self.times = times

    def load_track(self):
        self.workbook = openpyxl.load_workbook(self.track_file)
        self.worksheet = self.workbook[self.sheet]

    def start_race(self):
        self.load_track()
            
        for i in range(0,self.times):
            black = Bot("Black", self.worksheet)
            black.getMapValues(self.worksheet)


            while black.finish == False:
                black.drawCard()
                black.move()
                if black.finish== False:
                    black.getMapValues(self.worksheet)
            
            num_rows = 0
            for cell in self.worksheet['H']:
                if cell.value:
                    num_rows += 1

            value = black.turnstaken  # replace with the value you want to append
            print(f'Special turn required: {black.wassaveneeded} \n') if debug else None
            print(f'pasting turns taken to row {num_rows+1} in {self.sheet} [{self.worksheet.cell(row=2, column=7).value}]:  {black.turnstaken}')
            self.worksheet.cell(row=num_rows+1, column=8, value=value)
            self.workbook.save(self.track_file)


        
        

#MAIN

print(f'\npath\n{os.getcwd()}')
userinput = input('Debug? (y/n)')

if userinput == 'y':
    debug = True
else:
    debug = False

start = 1
end = 2
start = input('enter first map in range that you want to test. (i.e. 1,4,6)')
end = input('enter last map in range that you want to test. (i.e. 1,4,6)')
end = int(end)+1

input
raceArray = []
if debug:
    times = 1
else:
    times = 40
for i in range(int(start),int(end)):
    race = Race(f'{os.getcwd()}\RaceMaps.xlsx', f'MAP ({i})',times)
    raceArray.append(race)

for r in raceArray:
    r.start_race()

input('\nFINISHED press ENTER to quit')

#unfortanatly threading wont work because of reading and writing to a single file.
# threads = []
# for r in raceArray:
#     t = threading.Thread(target=r.start_race)
#     t.start()
#     threads.append(t)

# for t in threads:
#     t.join()

# This code defines a Bot class that has the attributes you specified. It also defines a Race class that takes the path to the Excel spreadsheet file and a list of bots as input.

# The Bot class has a move method that will be responsible for implementing the movement logic for each bot. The Race class has a load_track method that will parse the map data from the Excel spreadsheet and store it in a suitable format. The start method of the Race class will implement the logic for running the race.

# Note that this is just a skeleton and you'll need to fill in the details for each method to make it work correctly.